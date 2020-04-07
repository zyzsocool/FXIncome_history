import datetime
from dateutil.relativedelta import relativedelta
import openpyxl
from random import gauss
import numpy as np
import matplotlib.pyplot as plt
import copy
from tqdm import tqdm
import pandas as pd
plt.rcParams['font.sans-serif'] = ['SimHei']  # 用来正常显示中文标签
plt.rcParams['axes.unicode_minus'] = False  # 用来正常显示负号
pd.set_option('display.max_columns', None)
pd.set_option('display.width', 1000)
pd.set_option('display.float_format', lambda x: '%.0f' % x)
pd.set_option('display.unicode.ambiguous_as_wide', True)
pd.set_option('display.unicode.east_asian_width', True)


class FXIncome(object):
    def __init__(self):
        self.asset = []
        self.assetintial = []

        self.date = []
        self.daten = 0
        self.mu = []
        self.sigma = []

    def assetput(self, address):
        """读入债券资产的基础信息和操作信息、读入利率变化信息，并将其保存在类的属性里"""
        # 1读入债券资产信息和期间操作信息
        wb = openpyxl.load_workbook(address, data_only=True)
        ws = wb['债券']
        print('-----------------未入仓债券（有提前还款条款或浮动利率）---------------')
        for i in ws.rows:
            if i[0].value == '债券':  # 用来关键日期，包括开始日，各个操作日和最终日等
                self.date.append(i[11].value)

                ii = 12
                while i[ii].value != '结束':
                    self.date.append(i[ii].value)
                    ii += 1
                    self.daten += 1
            # 用来读债券要素（顺便算一下全价，如果算出来的全价和中债估值的不同，一般是因为这个债券是提前兑付本金的，这种债券在这个版本中不处理，后面版本再补）
            if i[0].value != '债券' and i[1].value is not None:
                bondname = i[0].value
                facevalue = float(i[1].value)
                coupondate = datetime.datetime.strptime(
                    i[3].value.replace('-', ''), '%Y%m%d')
                couponrate = float(i[4].value) / 100
                coupountype = i[5].value
                coupounfrequency = i[6].value
                maturity = float(i[7].value)
                ytm = float(i[8].value) / 100

                price = float(i[9].value) / 100
                enddate = datetime.datetime.strptime(
                    i[10].value.replace('-', ''), '%Y%m%d')

                ipv = self.pv(
                    coupondate,
                    couponrate,
                    coupountype,
                    coupounfrequency,
                    maturity,
                    ytm,
                    self.date[0])  # 算全价
                ipvcal = abs(price - ipv)
                if ipvcal < 0.001:  # 如果算出来的全价和中债估值一致，就入库

                    k = [
                        bondname,
                        facevalue,
                        coupondate,
                        couponrate,
                        coupountype,
                        coupounfrequency,
                        maturity,
                        ytm,
                        price,
                        enddate]
                    for kk in range(
                            11, 11 + self.daten):  # 在每个资产信息的末端，加入每只债券的操作信息
                        if i[kk].value:
                            amount = i[kk].value
                        else:
                            amount = 0

                        k.append(amount)
                    self.asset.append(k)
                else:  # 如果算出来的全价和中债估值不一致，就不入库，并且print出来进行提示

                    print(i[0].value)

        print('-----------------未入仓债券（有提前还款条款或浮动利率）---------------')

        self.assetintial = copy.deepcopy(
            self.asset)  # 后面会对asset进行不断操作，所以先复制一个保持不变。
        # 2.1读入变动利率法下的利率变化信息
        ws = wb['变动利率法']
        mu = []
        sigma = []
        for i in range(0, self.daten):
            mu.append(ws.cell(row=2, column=3 + i).value)
            sigma.append(ws.cell(row=3, column=3 + i).value / 1.96)
        self.mu = mu
        self.sigma = sigma

        mu_c = []
        sigma_c = []

        # 2.2读入利率曲线法下的利率变化信息
        ws = wb['利率曲线法']
        for i in range(0, self.daten + 1):
            mu_ct = []
            sigma_ct = []
            for j in range(0, 12):

                mu_ct.append(ws.cell(column=i + 2, row=j + 3).value)
                sigma_ct.append(ws.cell(column=i + 2, row=j + 18).value / 1.96)
            mu_c.append(mu_ct)
            sigma_c.append(sigma_ct)
        self.sigma_c = sigma_c
        self.mu_c = mu_c

    def pv(
            self,
            coupon_date,
            coupon_rate,
            coupon_type,
            coupon_frequency,
            maturity,
            ytm,
            assessment_date):
        """
        计算全价的函数，分三类债券讨论：附息债券、贴现债券和到期还本付息的债券
        :param coupon_date: 债券的起息日
        :param coupon_rate: 债券的票面利率
        :param coupon_type: 债券类型：分为附息、贴现和到期还本付息
        :param coupon_frequency: 每年支付利率次数
        :param maturity: 初始期限
        :param ytm: 到期收益率
        :param assessment_date: 核算日
        :return: 债券全价

        """

        # 1附息债券的计算方法
        if coupon_type == '附息':
            couponcount = 0
            while (
                    coupon_date -
                    assessment_date).days <= 0:  # 找到核算日在哪个付息区间内（就是找到上一付息日和下一付息日）
                lastpaydate = coupon_date
                coupon_date += relativedelta(months=12 / coupon_frequency)
                nextpaydate = coupon_date
                couponcount += 1
            d = (nextpaydate - assessment_date).days  # 离下一个付息日的天数
            ts = (nextpaydate - lastpaydate).days  # 所在付息区间的总天数
            if couponcount == maturity * coupon_frequency:  # 处于最后一个付息区间时的公式
                pv = (1 + coupon_rate / coupon_frequency) / \
                     (1 + ytm / coupon_frequency / ts * d)
            else:  # 非最后付息区间的公式
                dts = d / ts
                n = maturity * coupon_frequency - couponcount + 1
                pv = coupon_rate / coupon_frequency / (1 + ytm / coupon_frequency) ** (dts - 1) / (
                        ytm / coupon_frequency) * (1 - 1 / (1 + ytm / coupon_frequency) ** n) + 1 / (
                             1 + ytm / coupon_frequency) ** (dts + n - 1)
        # 2贴现债券的计算方法
        elif coupon_type == '贴现':
            if maturity < 0.26:  # 初始期限为3个月的贴现债券（按道理是0.25，但是因为约分的原因会有0.25+一点点的）
                maturedate = coupon_date + relativedelta(months=3)
            elif maturity < 0.55:  # 初始期限为6个月的贴现债券（按道理是0.5）
                maturedate = coupon_date + relativedelta(months=6)
            else:  # 一般贴现债券都不超过1年的，所以不是3个月和6个月的就是1年的
                maturedate = coupon_date + relativedelta(months=12)
            d = (maturedate - assessment_date).days
            # 贴现的PV就简单的，直接除就好了
            if d > 0:
                pv = 1 / (1 + ytm / 365 * d)
            else:
                pv = (1 - ytm / 365 * d)

        # 3到期一次还本付息债券的计算方法
        # 计算方法和贴现类似，不重复说明
        elif coupon_type == '到期一次还本付息':
            if maturity < 0.26:
                maturedate = coupon_date + relativedelta(months=3)
                couponcountday = (maturedate - coupon_date).days
            elif maturity < 0.55:
                maturedate = coupon_date + relativedelta(months=6)
                couponcountday = (maturedate - coupon_date).days
            else:
                maturedate = coupon_date + relativedelta(months=12)
                couponcountday = (maturedate - coupon_date).days
            d = (maturedate - assessment_date).days
            pv = (1 + coupon_rate * couponcountday / 365) / (1 + ytm / 365 * d)

        return pv

    def holdingprofit(
            self,
            coupondate,
            couponrate,
            coupountype,
            coupounfrequency,
            maturity,
            ytm,
            begindate,
            enddate,
            interestchange):
        """

        计算在持有期间，持有的某只债券在收益率变动特定量的情况下，获得的持有收益
        :param coupondate: 债券的起息日
        :param couponrate: 债券的票面利率
        :param coupountype: 债券的付息类型
        :param coupounfrequency: 债券一年付息次数
        :param maturity: 债券初始期限
        :param ytm: 债券的到期收益率
        :param begindate: 核算区间的起始日
        :param enddate: 核算区间的结束日
        :param interestchange: 核算区间内收益率的变化
        :return: 一个list:[起始日价值，结算日价值，核算区间内估值收益，核算区间内收到的利息]
        """

        # 1附息债券的算法
        if coupountype == '附息':
            bondenddate = coupondate + \
                relativedelta(months=12 * maturity)  # 先算债券的到期日
            if (bondenddate - begindate).days <= 0:  # 如果在开始日那天债券已经到期了，就假设按到期时的收益率进行再投资
                pv0 = 1 + ytm * (begindate - bondenddate).days / 365
                r = ytm + interestchange
                pv1 = 1 + r * (enddate - bondenddate).days / 365
                pvgain = pv1 - pv0
                interestgain = 0
            else:  # 如果在开始日那天债券还没有到期
                if (bondenddate - enddate).days >= 0:  # 如果在开始日那天债券还没有到期，且结束日那天也没有到期
                    pv0 = self.pv(
                        coupondate,
                        couponrate,
                        coupountype,
                        coupounfrequency,
                        maturity,
                        ytm,
                        begindate)
                    pv1 = self.pv(
                        coupondate,
                        couponrate,
                        coupountype,
                        coupounfrequency,
                        maturity,
                        ytm + interestchange,
                        enddate)

                    pvgain = pv1 - pv0
                    interestgain = -couponrate / coupounfrequency
                    while (coupondate - enddate).days <= 0:

                        coupondate += relativedelta(months=12 /
                                                    coupounfrequency)
                        if (coupondate - begindate).days > 0:
                            interestgain += couponrate / coupounfrequency
                else:  # 如果在开始日那天债券还没有到期，但是结束日那天已经到期了，就假设按到期时的收益率进行再投资
                    pv0 = self.pv(
                        coupondate,
                        couponrate,
                        coupountype,
                        coupounfrequency,
                        maturity,
                        ytm,
                        begindate)
                    r = ytm + interestchange
                    pv1 = 1 + r * (enddate - bondenddate).days / 365
                    pvgain = pv1 - pv0
                    interestgain = -couponrate / coupounfrequency
                    while (coupondate - bondenddate).days <= 0:
                        coupondate += relativedelta(months=12 /
                                                    coupounfrequency)
                        if (coupondate - begindate).days > 0:
                            interestgain += couponrate / coupounfrequency
        # 2贴现债券的算法
        elif coupountype == '贴现':
            if maturity < 0.26:
                bondenddate = coupondate + relativedelta(months=3)
            elif maturity < 0.55:
                bondenddate = coupondate + relativedelta(months=6)
            else:
                bondenddate = coupondate + relativedelta(months=12)
            if (bondenddate - enddate).days > 0:
                pv0 = self.pv(
                    coupondate,
                    couponrate,
                    coupountype,
                    coupounfrequency,
                    maturity,
                    ytm,
                    begindate)
                pv1 = self.pv(
                    coupondate,
                    couponrate,
                    coupountype,
                    coupounfrequency,
                    maturity,
                    ytm + interestchange,
                    enddate)

                pvgain = pv1 - pv0
                interestgain = 0
            else:
                pv0 = self.pv(
                    coupondate,
                    couponrate,
                    coupountype,
                    coupounfrequency,
                    maturity,
                    ytm,
                    begindate)
                r = ytm + interestchange
                pv1 = 1 + r * (enddate - bondenddate).days / 365
                pvgain = pv1 - pv0
                interestgain = 0
        # 3到期一次还本付息债券的算法
        elif coupountype == '到期一次还本付息':
            if maturity < 0.26:
                bondenddate = coupondate + relativedelta(months=3)
                couponcountday = (bondenddate - coupondate).days
            elif maturity < 0.55:
                bondenddate = coupondate + relativedelta(months=6)
                couponcountday = (bondenddate - coupondate).days
            else:
                bondenddate = coupondate + relativedelta(months=12)
                couponcountday = (bondenddate - coupondate).days

            if (bondenddate - begindate).days <= 0:
                pv0 = 1 + ytm * (begindate - bondenddate).days / 365
                r = ytm + interestchange
                pv1 = 1 + r * (enddate - bondenddate).days / 365
                pvgain = pv1 - pv0
                interestgain = 0
            else:
                if (bondenddate - enddate).days > 0:
                    pv0 = self.pv(
                        coupondate,
                        couponrate,
                        coupountype,
                        coupounfrequency,
                        maturity,
                        ytm,
                        begindate)
                    pv1 = self.pv(
                        coupondate,
                        couponrate,
                        coupountype,
                        coupounfrequency,
                        maturity,
                        ytm + interestchange,
                        enddate)

                    pvgain = pv1 - pv0
                    interestgain = 0
                else:
                    pv0 = self.pv(
                        coupondate,
                        couponrate,
                        coupountype,
                        coupounfrequency,
                        maturity,
                        ytm,
                        begindate)
                    r = ytm + interestchange
                    pv1 = 1 + r * (enddate - bondenddate).days / 365

                    pvgain = pv1 - pv0
                    interestgain = couponrate * couponcountday / 365

        return [pv0, pv1, pvgain, interestgain]

    def holdingprofitall(self, begindate, enddate, interestchange, im=None):
        """
        计算持仓债券在核算区间的收益，分为不操作的收益和操作的收益
        :param begindate: 核算区间的起始日
        :param enddate: 核算区间的结束日
        :param interestchange: 核算区间内到期收益率的变化
        :param im: 是一个指示量，没有值输入时是计算全区间的，有值输入时是计算分区间的
        :return: 输出一个list：[不操作情况下估值收益，不操作情况下利息收益，操作情况下估值收益，操作情况下利息收益，操作所造成的现金头寸缺口]

        注：这个函数是为了将整个核算区间分为多个小区间并进行收益核算所构建的，所以用这个函数之后会更改asset的属性（头寸和ytm）
            以便对区间进行连续计算，因此每次批量用这个函数前要对asset进行初始化。同时由于历史原因，这个函数只适用于变动利率法。
        """

        cash = 0
        pvgain0 = 0
        interestgain0 = 0
        pvgain = 0
        interestgain = 0
        for i in self.asset:
            m = self.holdingprofit(
                i[2],
                i[3],
                i[4],
                i[5],
                i[6],
                i[7],
                begindate,
                enddate,
                interestchange)
            pvgain0 += m[2] * i[1]
            interestgain0 += m[3] * i[1]
            if im is not None:

                i[1] += i[10 + im]
                cash -= i[10 + im] * m[0]
                i[7] += interestchange

            pvgain += m[2] * i[1]
            interestgain += m[3] * i[1]
        return([pvgain0 / 10000, interestgain0 / 10000, pvgain / 10000, interestgain / 10000, cash / 10000])

    def hdprofitplus(self, choice=0):
        """
        在变动利率法下计算区间收益
        holdingporfitall 计算小区间的收益，holdingprofitplus则是基于holdingporfitall上计算整个区间的收益

        :param choice: 一个指示量，输入变量是画图模式，不输入变量是计算模式
        :return: 画图模式下只输出一个list：[大区间下不操作的收益，区间下操作的收益]
                 计算模式下输出一个复杂一点的list[大区间下各个小区间的具体表现信息（DateFrame）,[大区间下不操作的收益，区间下操作的收益]]
                         同时将很多信息print出来
        """
        self.asset = copy.deepcopy(self.assetintial)
        mu = copy.deepcopy(self.mu)

        sigma = copy.deepcopy(self.sigma)
        if choice == 0:
            for i in range(0, len(sigma)):
                sigma[i] = 0
        interestchange = [0]
        for i in range(0, self.daten):
            change = gauss(mu[i], sigma[i])
            interestchange[0] += change
            interestchange.append(change)

        result = {}
        result0 = self.holdingprofitall(
            self.date[0], self.date[-1], interestchange[0])

        result[self.date[0].strftime(
            '%Y%m%d') + '---' + self.date[-1].strftime('%Y%m%d')] = result0
        holding = result0[0] + result0[1]

        changeprofitpv = 0
        changeprofitin = 0
        cashn = 0
        for i in range(0, self.daten):
            result0 = self.holdingprofitall(
                self.date[i], self.date[i + 1], interestchange[i + 1], i)
            changeprofitpv += result0[2]
            changeprofitin += result0[3]
            cashn += result0[4]

            result[self.date[i].strftime(
                '%Y%m%d') + '---' + self.date[i + 1].strftime('%Y%m%d')] = result0

        changeprofit = changeprofitpv + changeprofitin
        result[self.date[0].strftime(
            '%Y%m%d') + '---' + self.date[- 1].strftime('%Y%m%d')][2] = changeprofitpv
        result[self.date[0].strftime(
            '%Y%m%d') + '---' + self.date[- 1].strftime('%Y%m%d')][3] = changeprofitin
        result[self.date[0].strftime(
            '%Y%m%d') + '---' + self.date[- 1].strftime('%Y%m%d')][4] = cashn
        resultpd = pd.DataFrame.from_dict(
            result,
            orient='index',
            columns=[
                '估值收益（不操作）',
                '利息收益（不操作）',
                '估值收益（操作）',
                '利息收益（操作）',
                '现金头寸'])
        if choice != 0:
            return [holding, changeprofit]
        else:
            print(resultpd)

            print('不操作收益(万元)：%.0f' % holding)
            print('操作收益（万元）：%.0f' % changeprofit)
            print('操作净收益（万元）：%.0f' % (changeprofit - holding))
            print('收益增加率（万元）：%.0f' % (changeprofit - holding) / holding)
            return [resultpd, [holding, changeprofit]]

    def hdprofitplus2(self, choice=0):
        """
        利率曲线法下计算区间收益
        注：和 hdprofitplus不同，hdprofitplus2直接调用holdingprofit函数，没有用到holdingprofitall函数

        :param choice: 一个指示量，输入变量是画图模式，不输入变量是计算模式
        :return: 画图模式下只输出一个list：[大区间下不操作的收益，区间下操作的收益]
                 计算模式下输出一个复杂一点的list[大区间下各个小区间的具体表现信息（DateFrame）,[大区间下不操作的收益，区间下操作的收益]]
                         同时将很多信息print出来
        """
        self.asset = copy.deepcopy(self.assetintial)
        mu = copy.deepcopy(self.mu_c)

        sigma = copy.deepcopy(self.sigma_c)

        interest = []
        if choice != 0:

            for i in range(0, self.daten + 1):
                interest_t = []
                for j in range(0, 12):
                    interest_t.append(gauss(mu[i][j], sigma[i][j]))
                interest.append(interest_t)
        else:
            interest = mu

        result = {}

        changeprofitpv = 0
        changeprofitin = 0
        cashn = 0
        pv0 = 0

        for i in range(0, self.daten + 1):
            cash = 0
            pvgain0 = 0
            interestgain0 = 0
            pvgain = 0
            interestgain = 0

            if i == 0:
                ik1 = 0
                ik2 = self.daten - 1
                ik0 = 0
            else:
                ik1 = i - 1
                ik2 = i - 1
                ik0 = 1
            for asset in self.asset:

                daylastbegin = (asset[9] - self.date[ik1]).days / 365
                daylastend = (asset[9] - self.date[ik2 + 1]).days / 365
                if daylastbegin <= 0:
                    ytmbegin = interest[ik1][0]

                elif daylastbegin <= 0.25:
                    ytmbegin = interest[ik1][0] + (
                        interest[ik1][1] - interest[ik1][0]) / 0.25 * daylastbegin
                elif daylastbegin <= 0.5:
                    ytmbegin = interest[ik1][1] + (
                        interest[ik1][2] - interest[ik1][1]) / 0.25 * (daylastbegin - 0.25)
                elif daylastbegin <= 0.75:
                    ytmbegin = interest[ik1][2] + (
                        interest[ik1][3] - interest[ik1][2]) / 0.25 * (daylastbegin - 0.5)
                elif daylastbegin <= 1:
                    ytmbegin = interest[ik1][3] + (
                        interest[ik1][4] - interest[ik1][3]) / 0.25 * (daylastbegin - 0.75)
                elif daylastbegin <= 2:
                    ytmbegin = interest[ik1][4] + \
                        (interest[ik1][5] - interest[ik1][4]) / 1 * (daylastbegin - 1)
                elif daylastbegin <= 3:
                    ytmbegin = interest[ik1][5] + \
                        (interest[ik1][6] - interest[ik1][5]) / 1 * (daylastbegin - 2)
                elif daylastbegin <= 4:
                    ytmbegin = interest[ik1][6] + \
                        (interest[ik1][7] - interest[ik1][6]) / 1 * (daylastbegin - 3)
                elif daylastbegin <= 5:
                    ytmbegin = interest[ik1][7] + \
                        (interest[ik1][8] - interest[ik1][7]) / 1 * (daylastbegin - 4)
                elif daylastbegin <= 10:
                    ytmbegin = interest[ik1][8] + \
                        (interest[ik1][9] - interest[ik1][8]) / 5 * (daylastbegin - 5)
                elif daylastbegin <= 20:
                    ytmbegin = interest[ik1][9] + (
                        interest[ik1][10] - interest[ik1][9]) / 10 * (daylastbegin - 10)
                elif daylastbegin <= 30:
                    ytmbegin = interest[ik1][10] + (
                        interest[ik1][11] - interest[ik1][10]) / 10 * (daylastbegin - 20)

                if daylastend <= 0:
                    ytmend = interest[ik2 + 1][0]
                elif daylastend <= 0.25:
                    ytmend = interest[ik2 + 1][0] + \
                        (interest[ik2 + 1][1] - interest[ik2 + 1][0]) / 0.25 * daylastend
                elif daylastend <= 0.5:
                    ytmend = interest[ik2 + 1][1] + (
                        interest[ik2 + 1][2] - interest[ik2 + 1][1]) / 0.25 * (daylastend - 0.25)
                elif daylastend <= 0.75:
                    ytmend = interest[ik2 + 1][2] + (
                        interest[ik2 + 1][3] - interest[ik2 + 1][2]) / 0.25 * (daylastend - 0.5)
                elif daylastend <= 1:
                    ytmend = interest[ik2 + 1][3] + (
                        interest[ik2 + 1][4] - interest[ik2 + 1][3]) / 0.25 * (daylastend - 0.75)
                elif daylastend <= 2:
                    ytmend = interest[ik2 + 1][4] + (
                        interest[ik2 + 1][5] - interest[ik2 + 1][4]) / 1 * (daylastend - 1)
                elif daylastend <= 3:
                    ytmend = interest[ik2 + 1][5] + (
                        interest[ik2 + 1][6] - interest[ik2 + 1][5]) / 1 * (daylastend - 2)
                elif daylastend <= 4:
                    ytmend = interest[ik2 + 1][6] + (
                        interest[ik2 + 1][7] - interest[ik2 + 1][6]) / 1 * (daylastend - 3)
                elif daylastend <= 5:
                    ytmend = interest[ik2 + 1][7] + (
                        interest[ik2 + 1][8] - interest[ik2 + 1][7]) / 1 * (daylastend - 4)
                elif daylastend <= 10:
                    ytmend = interest[ik2 + 1][8] + (
                        interest[ik2 + 1][9] - interest[ik2 + 1][8]) / 5 * (daylastend - 5)
                elif daylastend <= 20:
                    ytmend = interest[ik2 + 1][9] + (
                        interest[ik2 + 1][10] - interest[ik2 + 1][9]) / 10 * (daylastend - 10)
                elif daylastend <= 30:
                    ytmend = interest[ik2 + 1][10] + (
                        interest[ik2 + 1][11] - interest[ik2 + 1][10]) / 10 * (daylastend - 20)

                interestchange = ytmend - ytmbegin
                m = self.holdingprofit(asset[2],
                                       asset[3],
                                       asset[4],
                                       asset[5],
                                       asset[6],
                                       ytmbegin,
                                       self.date[ik1],
                                       self.date[ik2 + 1],
                                       interestchange)
                pvgain0 += m[2] * asset[1]
                interestgain0 += m[3] * asset[1]

                asset[1] += asset[10 + ik1] * ik0
                cash -= asset[10 + ik1] * m[0] * ik0

                asset[7] = ytmbegin

                pvgain += m[2] * asset[1]
                interestgain += m[3] * asset[1]
                if i == 0:
                    pv0 += m[0] * asset[1] / 10000
            if i == 0:
                result0 = [
                    pvgain0 /
                    10000,
                    interestgain0 /
                    10000,
                    None,
                    None,
                    None]

            else:
                result0 = [
                    pvgain0 / 10000,
                    interestgain0 / 10000,
                    pvgain / 10000,
                    interestgain / 10000,
                    cash / 10000]
            result[self.date[ik1].strftime(
                '%Y%m%d') + '---' + self.date[ik2 + 1].strftime('%Y%m%d')] = result0

            if i == 0:
                holding = result0[0] + result0[1]

            else:
                changeprofitpv += result0[2]
                changeprofitin += result0[3]
                cashn += result0[4]

        changeprofit = changeprofitpv + changeprofitin
        result[self.date[0].strftime(
            '%Y%m%d') + '---' + self.date[-1].strftime('%Y%m%d')][2] = changeprofitpv
        result[self.date[0].strftime(
            '%Y%m%d') + '---' + self.date[-1].strftime('%Y%m%d')][3] = changeprofitin
        result[self.date[0].strftime(
            '%Y%m%d') + '---' + self.date[-1].strftime('%Y%m%d')][4] = cashn
        resultpd = pd.DataFrame.from_dict(
            result,
            orient='index',
            columns=[
                '估值收益（不操作）',
                '利息收益（不操作）',
                '估值收益（操作）',
                '利息收益（操作）',
                '现金头寸'])
        if choice != 0:
            return [holding, changeprofit]
        else:

            print(resultpd)
            dayratio = 365 / (self.date[-1] - self.date[0]).days
            print('初始持仓估值(万元)：%.0f' % pv0)
            print('不操作收益(万元)：%.0f' % holding)
            print('不操作年化收益：%.2f%%' % (100 * holding / pv0 * dayratio))
            print('操作收益（万元）：%.0f' % changeprofit)
            print('操作年化收益：%.2f%%' % (100 * changeprofit / pv0 * dayratio))
            print('操作净收益（万元）：%.0f' % (changeprofit - holding))

            print('操作年化净收益：%.2f%%' %
                  (100 * (changeprofit - holding) / pv0 * dayratio))
            return [resultpd, [holding, changeprofit]]

    def hdprofitplot(self, choice, test=10000):
        """

        这个函数是进行多次模拟（hdprofitplus或hdprofitplus2）并画图的函数
        :param choice: 输入1就用变动利率法画图，输入2就用利率曲线法画图
        :param test: 不输入参数的话默认模拟是10000次，输入参数就按输入的次数模拟
        :return:这个不返回函数，直接出图
        """
        # 1先用统计方法调用一次hdprofitplus或hdprofitplus2（可以计算均值在图上画出来，顺便print出信息）
        self.asset = copy.deepcopy(self.assetintial)
        if choice == 1:
            profit = self.hdprofitplus()
        elif choice == 2:
            profit = self.hdprofitplus2()
        profithonding = int(profit[1][0])
        profitchange = int(profit[1][1])
        profitdiv = profitchange - profithonding
        # 2用画图模式多次调用hdprofitplus或hdprofitplus2

        holding = []
        changeprofit = []
        divprofit = []
        for i in tqdm(range(0, test)):
            self.asset = copy.deepcopy(self.assetintial)
            if choice == 1:

                profit = self.hdprofitplus(1)
            elif choice == 2:
                profit = self.hdprofitplus2(1)
            div = profit[1] - profit[0]
            holding.append(profit[0])
            changeprofit.append(profit[1])
            divprofit.append(div)

        holding = np.array(holding)
        changeprofit = np.array(changeprofit)

        date0 = self.date[0].strftime('%Y%m%d')
        date1 = self.date[-1].strftime('%Y%m%d')

        figure = plt.figure(1)
        figure.suptitle('持有期收益测试%s--%s（实验次数：%d）' % (date0, date1, test))

        ax1 = plt.subplot(1, 2, 1)
        ax2 = plt.subplot(1, 2, 2)

        plt.sca(ax1)
        plt.hist(holding, histtype='stepfilled', alpha=0.3, density=1,
                 bins=200, label='不操作(均值:' + str(profithonding) + ')')
        plt.hist(
            changeprofit,
            histtype='stepfilled',
            alpha=0.3,
            density=1,
            bins=200,
            label='操作(均值:' + str(profitchange) + ')')
        plt.xlabel('损益（估值+利息）(万元)')
        plt.title('损益相对频率图')

        plt.ylabel('相对频率')
        plt.legend(loc=1)

        plt.sca(ax2)
        plt.hist(
            divprofit,
            histtype='stepfilled',
            alpha=0.3,
            density=1,
            bins=200,
            color='green',
            label='操作与不操作之差(均值:' + str(profitdiv) + ')')
        plt.legend(loc=1)
        plt.xlabel('损益差（估值+利息）(万元)')
        plt.ylabel('相对频率')
        plt.title('损益差相对频率图')

        plt.show()


b = FXIncome()
address = './计算表.xlsx'
b.assetput(address)

b.hdprofitplot(2, 100)
